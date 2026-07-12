# api/registro_diario.py
from fastapi import APIRouter, Depends, HTTPException, Form, File, UploadFile
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime
import os
import uuid

from db.database import get_db
from db.models import User, RegistroDiario, Area
from schemas.registro_diario_schema import (
    RegistroDiarioCreate,
    RegistroDiarioResponse,
    RegistroDiarioCalidadUpdate,
    RegistroDiarioOperacionesUpdate,
)
from api.deps import get_current_user

import io
from datetime import date
from typing import Optional
from fastapi import Query
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from openpyxl.drawing.image import Image as OpenpyxlImage

from services.ftp_service import upload_image_bytes, download_image_bytes
from PIL import Image as PILImage

router = APIRouter(prefix="/api/registros-diarios", tags=["Registros Diarios"])

AREA_CALIDAD_ID = 25
AREA_OPERACIONES_ID = 26

# Carpeta donde se guardan las imágenes subidas
UPLOADS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads")
os.makedirs(UPLOADS_DIR, exist_ok=True)


TIPOS_ACTIVIDAD_ENTREGABLE = [
    "Desarrollo de Proyecto",
    "Informe",
    "Entregable Final",
]

_PANEL_COLUMNS = (
    # ── Identificadores y metadatos ──
    RegistroDiario.id,
    RegistroDiario.usuario_id,
    RegistroDiario.area_id,
    RegistroDiario.fecha_registro,
    Area.nombre.label("area_nombre"),
    User.name.label("trabajador_nombre"),
    # ── Campos del colaborador ──
    RegistroDiario.proceso,
    RegistroDiario.tipo_actividad,
    RegistroDiario.tipo_tarea,
    RegistroDiario.entregable,
    RegistroDiario.responsable_asigna,
    RegistroDiario.fecha_inicio,
    RegistroDiario.fecha_entrega,
    RegistroDiario.unidad_medida,
    RegistroDiario.tiempo_estimado,
    RegistroDiario.estado_base,
    # ── Campos de Calidad ──
    RegistroDiario.estado_entregable_calidad,
    RegistroDiario.estado_animo,
    RegistroDiario.observaciones_calidad,
    RegistroDiario.tiempo_estandar,
    RegistroDiario.tiempo_real_calidad,
    RegistroDiario.errores_observaciones,
    RegistroDiario.eficiencia,
    RegistroDiario.tasa_calidad,
    RegistroDiario.rubrica_final,
    RegistroDiario.auditado_calidad,
    RegistroDiario.prioridad,
    RegistroDiario.tiempo_real_operaciones,
    RegistroDiario.estado_tarea_operaciones,
    RegistroDiario.motivo_retraso,
    RegistroDiario.observaciones_operaciones,
    RegistroDiario.enlace_evidencia,
    RegistroDiario.imagen_evidencia,
    RegistroDiario.validacion_lider,
    RegistroDiario.actitud_colaborador,
    RegistroDiario.dias_vencimiento,
    RegistroDiario.auditado_operaciones,
)

@router.post("/", response_model=RegistroDiarioResponse)
def crear_registro(
    registro_in: RegistroDiarioCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Verificación de seguridad básica
    if not current_user.kpi_area_id:
        raise HTTPException(status_code=400, detail="El usuario no tiene un área asignada.")

    # Inyección de campos automáticos
    nuevo_registro = RegistroDiario(
        usuario_id=current_user.id,
        area_id=current_user.kpi_area_id,
        fecha_registro=datetime.utcnow(),
        # Expansión de los datos obligatorios del usuario
        **registro_in.dict()
    )
    
    db.add(nuevo_registro)
    db.commit()
    db.refresh(nuevo_registro)
    
    return nuevo_registro


# ══════════════════════════════════════════════════════════════════════════
#  PANELES DE LECTURA (Operaciones / Calidad)
#  Solo lectura: aquí no se edita nada todavía, solo se lista lo que el
#  colaborador ya registró, filtrando columnas y filas según el área.
# ══════════════════════════════════════════════════════════════════════════

@router.get("/panel-operaciones", response_model=List[RegistroDiarioResponse])
def panel_operaciones(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Panel de Operaciones: ve TODAS las actividades registradas
    (reuniones, charlas, gestión administrativa, entregables, etc.)
    Solo accesible por usuarios cuya área es Operaciones (id 26).
    """
    if current_user.kpi_area_id != AREA_OPERACIONES_ID:
        raise HTTPException(status_code=403, detail="No tienes acceso a este panel.")

    registros = (
        db.query(*_PANEL_COLUMNS)
        .join(User, RegistroDiario.usuario_id == User.id)
        .join(Area, RegistroDiario.area_id == Area.id)
        .order_by(RegistroDiario.fecha_registro.desc())
        .all()
    )
    return registros


@router.get("/panel-calidad", response_model=List[RegistroDiarioResponse])
def panel_calidad(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Panel de Calidad: ve SOLO las actividades de tipo "entregable"
    (Desarrollo de Proyecto, Informe, Entregable Final), ya que es lo
    único que Calidad debe evaluar.
    Solo accesible por usuarios cuya área es Gestión de calidad (id 25).
    """
    if current_user.kpi_area_id != AREA_CALIDAD_ID:
        raise HTTPException(status_code=403, detail="No tienes acceso a este panel.")

    registros = (
        db.query(*_PANEL_COLUMNS)
        .join(User, RegistroDiario.usuario_id == User.id)
        .join(Area, RegistroDiario.area_id == Area.id)
        .order_by(RegistroDiario.fecha_registro.desc())
        .all()
    )
    return registros


# ══════════════════════════════════════════════════════════════════════════
#  EXPORTACIÓN A EXCEL
# ══════════════════════════════════════════════════════════════════════════

COLUMNAS_CALIDAD = [
    "Fecha Registro", "Área", "Trabajador", "Proceso", "Tipo de Actividad",
    "Entregable Específico", "Estado del Entregable", "Fecha de Inicio",
    "Fecha Límite", "Estado de Ánimo", "Observaciones de Calidad",
    "Tiempo Estándar", "Tiempo Real", "Unidades (Fijo: 1)",
    "Errores Encontrados", "Eficiencia", "Tasa de Calidad", "Rúbrica Final",
]

COLUMNAS_OPERACIONES = [
    "Fecha Registro", "Área", "Trabajador", "Responsable que asigna",
    "Entregable Específico", "Tipo de Tarea", "Prioridad", "Unidad de Medida",
    "Tiempo Estimado", "Fecha Límite", "Tiempo Real Operaciones",
    "Estado de Tarea", "Motivo Retraso", "Observaciones Operaciones",
    "Enlace Evidencia", "Validación Líder", "Actitud Colaborador",
    "Días Vencimiento",
]


def _fmt_fecha(valor):
    """Formatea datetime/date a string legible, o devuelve '' si es None."""
    if valor is None:
        return ""
    if hasattr(valor, "strftime"):
        return valor.strftime("%Y-%m-%d %H:%M") if hasattr(valor, "hour") else valor.strftime("%Y-%m-%d")
    return valor


def _fila_calidad(r) -> list:
    return [
        _fmt_fecha(r.fecha_registro),
        r.area_nombre,
        r.trabajador_nombre,
        r.proceso,
        r.tipo_actividad,
        r.entregable,
        r.estado_entregable_calidad,
        _fmt_fecha(r.fecha_inicio),
        _fmt_fecha(r.fecha_entrega),
        r.estado_animo,
        r.observaciones_calidad,
        r.tiempo_estandar,
        r.tiempo_real_calidad,
        1,
        r.errores_observaciones,
        r.eficiencia,
        r.tasa_calidad,
        r.rubrica_final,
    ]


def _fila_operaciones(r) -> list:
    return [
        _fmt_fecha(r.fecha_registro),
        r.area_nombre,
        r.trabajador_nombre,
        r.responsable_asigna,
        r.entregable,
        r.tipo_tarea,
        r.prioridad,
        r.unidad_medida,
        r.tiempo_estimado,
        _fmt_fecha(r.fecha_entrega),
        r.tiempo_real_operaciones,
        r.estado_tarea_operaciones,
        r.motivo_retraso,
        r.observaciones_operaciones,
        r.enlace_evidencia,
        r.validacion_lider,
        r.actitud_colaborador,
        r.dias_vencimiento,
    ]


def _parse_fecha_query(valor: Optional[str]) -> Optional[date]:
    """Convierte un query param string a date. Devuelve None si viene vacío o inválido."""
    if not valor or not valor.strip():
        return None
    try:
        return date.fromisoformat(valor.strip())
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Formato de fecha inválido: '{valor}'. Use YYYY-MM-DD.",
        )


@router.get("/exportar-excel")
def exportar_excel(
    area_panel: str = Query(..., description="'calidad' u 'operaciones'"),
    fecha: Optional[str] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    area_filtro: Optional[str] = Query(None),
    trabajador: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Exporta a Excel (.xlsx) los registros del panel de Calidad u Operaciones,
    aplicando los mismos filtros que se ven en pantalla.
    """
    area_panel = area_panel.lower().strip()
    if area_panel not in ("calidad", "operaciones"):
        raise HTTPException(status_code=400, detail="area_panel debe ser 'calidad' u 'operaciones'.")

    # ── Verificación de permisos (igual que en los paneles de lectura) ──
    if area_panel == "calidad" and current_user.kpi_area_id != AREA_CALIDAD_ID:
        raise HTTPException(status_code=403, detail="No tienes acceso a este panel.")
    if area_panel == "operaciones" and current_user.kpi_area_id != AREA_OPERACIONES_ID:
        raise HTTPException(status_code=403, detail="No tienes acceso a este panel.")

    # ── Parseo seguro de fechas (strings vacíos → None, no rompe validación) ──
    fecha_parsed = _parse_fecha_query(fecha)
    fecha_desde_parsed = _parse_fecha_query(fecha_desde)
    fecha_hasta_parsed = _parse_fecha_query(fecha_hasta)

    # Limpieza de strings vacíos en los demás filtros (para que "" no filtre por error)
    area_filtro = area_filtro.strip() if area_filtro and area_filtro.strip() else None
    trabajador = trabajador.strip() if trabajador and trabajador.strip() else None
    estado = estado.strip() if estado and estado.strip() else None

    # ── Construcción del query base (idéntico a los paneles existentes) ──
    query = (
        db.query(*_PANEL_COLUMNS)
        .join(User, RegistroDiario.usuario_id == User.id)
        .join(Area, RegistroDiario.area_id == Area.id)
    )

    # Solo se exportan registros YA AUDITADOS por el área correspondiente.
    # (Antes se exportaba todo, incluyendo lo que el colaborador acababa
    # de registrar y aún no había pasado por auditoría.)
    if area_panel == "calidad":
        query = query.filter(RegistroDiario.auditado_calidad == True)
    else:
        query = query.filter(RegistroDiario.auditado_operaciones == True)

    # ── Filtros comunes por query params ──
    if fecha_parsed:
        query = query.filter(db.func.date(RegistroDiario.fecha_registro) == fecha_parsed)

    if fecha_desde_parsed:
        query = query.filter(RegistroDiario.fecha_registro >= fecha_desde_parsed)

    if fecha_hasta_parsed:
        query = query.filter(RegistroDiario.fecha_registro <= fecha_hasta_parsed)

    if area_filtro:
        query = query.filter(Area.nombre.ilike(f"%{area_filtro}%"))

    if trabajador:
        query = query.filter(User.name.ilike(f"%{trabajador}%"))

    if estado:
        if area_panel == "calidad":
            query = query.filter(RegistroDiario.estado_entregable_calidad == estado)
        else:
            query = query.filter(RegistroDiario.estado_tarea_operaciones == estado)

    registros = query.order_by(RegistroDiario.fecha_registro.desc()).all()

    # ── Generación del archivo Excel en memoria ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calidad" if area_panel == "calidad" else "Operaciones"

    columnas = COLUMNAS_CALIDAD if area_panel == "calidad" else COLUMNAS_OPERACIONES
    fila_fn = _fila_calidad if area_panel == "calidad" else _fila_operaciones

    # Encabezados
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, titulo in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=titulo)
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal="center", vertical="center")

    # Filas de datos
    for row_idx, registro in enumerate(registros, start=2):
        for col_idx, valor in enumerate(fila_fn(registro), start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
        
        # Ajustamos altura de fila para que la imagen quepa visualmente
        ws.row_dimensions[row_idx].height = 80
        
        # Si es Operaciones e incrustó una imagen física
        if area_panel == "operaciones" and registro.imagen_evidencia:
            filename = registro.imagen_evidencia.split("/")[-1]
            try:
                contenido_img = download_image_bytes(filename)
                pil_img = PILImage.open(io.BytesIO(contenido_img))
                img = OpenpyxlImage(pil_img)
                img.width = 100
                img.height = 100
                col_letra = get_column_letter(15)
                ws.add_image(img, f"{col_letra}{row_idx}")
                ws.cell(row=row_idx, column=15, value="(Ver imagen adjunta)")
            except Exception as e:
                print(f"Error cargando imagen en Excel: {e}")

    # Ancho de columnas automático (aprox.)
    for col_idx, titulo in enumerate(columnas, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(titulo) + 2, 15)

    ws.freeze_panes = "A2"

    # ── Volcado a BytesIO y respuesta streaming ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"registros_{area_panel}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'},
    )


@router.get("/{registro_id}", response_model=RegistroDiarioResponse)
def obtener_registro_detalle(
    registro_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Trae el detalle completo de un registro diario específico.
    Se usa para mostrar el Modal antes de realizar la auditoría.
    """
    registro = db.query(RegistroDiario).filter(RegistroDiario.id == registro_id).first()
    if not registro:
        raise HTTPException(status_code=404, detail="Registro no encontrado")
    registro.trabajador_nombre = registro.usuario.name if registro.usuario else "Usuario Desconocido"
    registro.area_nombre = registro.area.nombre if registro.area else "Área Desconocida"
    
    return registro


@router.patch("/{registro_id}/calidad", response_model=RegistroDiarioResponse)
def auditar_calidad(
    registro_id: int,
    payload: RegistroDiarioCalidadUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Guarda la auditoría realizada por el área de Gestión de Calidad.
    """
    # Verificación de seguridad (Solo Calidad puede auditar aquí)
    if current_user.kpi_area_id != AREA_CALIDAD_ID:
        raise HTTPException(status_code=403, detail="No tienes permisos para auditar en Calidad.")

    registro = db.query(RegistroDiario).filter(RegistroDiario.id == registro_id).first()
    if not registro:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    # Actualizar los campos exclusivos de Calidad
    registro.estado_entregable_calidad = payload.estado_entregable_calidad
    registro.estado_animo = payload.estado_animo
    registro.tiempo_estandar = payload.tiempo_estandar
    registro.tiempo_real_calidad = payload.tiempo_real_calidad
    registro.errores_observaciones = payload.errores_observaciones
    registro.observaciones_calidad = payload.observaciones_calidad
    registro.rubrica_final = payload.rubrica_final
    registro.eficiencia = payload.eficiencia
    registro.tasa_calidad = payload.tasa_calidad
    
    # ¡Importante! Cambiar la bandera a True para que pase al historial
    registro.auditado_calidad = True

    db.commit()
    db.refresh(registro)
    
    # Inyectamos los nombres para el response model
    registro.trabajador_nombre = registro.usuario.name if registro.usuario else "Usuario Desconocido"
    registro.area_nombre = registro.area.nombre if registro.area else "Área Desconocida"

    return registro


@router.patch("/{registro_id}/operaciones", response_model=RegistroDiarioResponse)
async def auditar_operaciones(
    registro_id: int,
    prioridad: str = Form(""),
    tiempo_real_operaciones: float = Form(0.0),
    estado_tarea_operaciones: str = Form(""),
    motivo_retraso: str = Form(""),
    actitud_colaborador: str = Form(""),
    enlace_evidencia: str = Form(""),
    validacion_lider: str = Form(""),
    observaciones_operaciones: str = Form(""),
    dias_vencimiento: int = Form(0),
    imagen_evidencia: UploadFile = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    unidad_medida: str = Form("Horas"),
    tiempo_estimado: float = Form(0.0),
):
    """
    Guarda la auditoría realizada por el área de Operaciones.
    Acepta multipart/form-data para permitir subida de imagen.
    """
    if current_user.kpi_area_id != AREA_OPERACIONES_ID:
        raise HTTPException(status_code=403, detail="No tienes permisos para auditar en Operaciones.")

    registro = db.query(RegistroDiario).filter(RegistroDiario.id == registro_id).first()
    if not registro:
        raise HTTPException(status_code=404, detail="Registro no encontrado")

    # Actualizar los campos exclusivos de Operaciones
    registro.prioridad = prioridad
    registro.tiempo_real_operaciones = tiempo_real_operaciones
    registro.estado_tarea_operaciones = estado_tarea_operaciones
    registro.motivo_retraso = motivo_retraso
    registro.actitud_colaborador = actitud_colaborador
    registro.enlace_evidencia = enlace_evidencia
    registro.validacion_lider = validacion_lider
    registro.observaciones_operaciones = observaciones_operaciones
    registro.dias_vencimiento = dias_vencimiento
    registro.unidad_medida = unidad_medida
    registro.tiempo_estimado = tiempo_estimado

    # Guardar imagen si viene en la petición
    if imagen_evidencia and imagen_evidencia.filename:
        extension = os.path.splitext(imagen_evidencia.filename)[1]
        nombre_unico = f"{uuid.uuid4().hex}{extension}"
        contenido = await imagen_evidencia.read()
        url_publica = upload_image_bytes(contenido, nombre_unico)
        registro.imagen_evidencia = url_publica

    # Cambiar la bandera a True para que pase al historial
    registro.auditado_operaciones = True

    db.commit()
    db.refresh(registro)
    
    # Inyectamos los nombres para el response model
    registro.trabajador_nombre = registro.usuario.name if registro.usuario else "Usuario Desconocido"
    registro.area_nombre = registro.area.nombre if registro.area else "Área Desconocida"

    return registro