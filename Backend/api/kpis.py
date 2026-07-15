import os, re
import openpyxl
import pandas as pd
import time as _time
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import text
from typing import List, Optional
from datetime import datetime

from db.database import get_db
from db.models import User, Kpi, RegistroKpi, KpiCampo, RegistroValores, Area, KpiProgramado
from schemas.kpi_schema import KpiResponse, RegistroCreate, KpiProgramar
from api.deps import get_current_user
from services.notification_service import crear_notificacion

from services.cache_service import get_cache, set_cache, invalidate_cache_prefix

router = APIRouter(prefix="/api/kpis", tags=["Gestión de KPIs"])

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES (sin cambios)
# ══════════════════════════════════════════════════════════════════════════════

def _safe_float(val):
    if val is None or str(val).strip() == "":
        return None
        
    val_str = str(val).strip()
    
    # MAGIA: Si el string es una fecha YYYY-MM-DD, lo convertimos a días (como Excel/JS)
    if re.match(r"^\d{4}-\d{2}-\d{2}", val_str):
        try:
            # Tomamos solo la parte YYYY-MM-DD
            fecha = datetime.strptime(val_str[:10], "%Y-%m-%d")
            epoch = datetime(1970, 1, 1) # Misma fecha base que JavaScript
            return (fecha - epoch).total_seconds() / 86400.0
        except ValueError:
            pass

    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _campo_key(label: str) -> str:
    return re.sub(r"[^a-z0-9_]", "_", label.lower())[:80]


def is_system_col(h: str) -> bool:
    h_lower = h.lower().strip()
    exact_matches = {
        "semana", "cuartil", "fecha inicio", "fecha fin", "duración (días)",
        "kpi", "fórmula base", "objetivo", "función", "importancia", "responsable",
    }
    return h_lower in exact_matches


def build_js_formula(formula_text, campos_entrada):
    if not formula_text or str(formula_text).strip().lower() == "nan":
        return ""
    form = str(formula_text)
    form = re.sub(r"(?i)(?:\*|x|×)\s*100\b", "", form).strip()
    form = form.replace("×", "*").replace("x", "*").replace("÷", "/").replace("−", "-")
    campos_validos = [
        c for c in campos_entrada
        if c.lower() not in ["observaciones", "acciones correctivas"]
    ]
    campos_validos.sort(key=len, reverse=True)
    for c in campos_validos:
        form = form.replace(c, f"[{c}]")
    return form


def _extract_kpi_fields_from_sheet(ws) -> dict:
    rows = list(ws.iter_rows(min_row=4, max_row=5, values_only=True))
    if len(rows) < 2:
        return {}
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    data = rows[1]

    def get_val(col_name):
        try:
            idx = headers.index(col_name)
            v = data[idx] if idx < len(data) else None
            return None if v is None or str(v).strip() == "nan" else v
        except ValueError:
            return None

    kpi_name = get_val("KPI")
    if not kpi_name or str(kpi_name).strip() in ("nan", ""):
        kpi_name = None

    formula = get_val("Fórmula base") or ""
    meta_raw = (
        get_val("Meta KPI <=") or get_val("Meta KPI >=")
        or get_val("Meta KPI ≤") or get_val("Meta KPI ≥")
    )
    meta_tipo = "minimo"
    for h in headers:
        if "Meta KPI <" in h or "Meta KPI ≤" in h:
            meta_tipo = "maximo"
            break

    meta_prod_raw = (
        get_val("Meta Producción <=") or get_val("Meta Producción >=")
        or get_val("Meta Producción ≤") or get_val("Meta Producción ≥")
    )
    horas_plan_raw = get_val("Horas planificadas")

    input_fields, seen = [], set()
    for h in headers:
        h_clean = h.strip()
        if h_clean and not is_system_col(h_clean) and h_clean not in seen:
            input_fields.append(h_clean)
            seen.add(h_clean)

    return {
        "kpi_name": str(kpi_name) if kpi_name else None,
        "formula": str(formula),
        "meta_valor": _safe_float(meta_raw),
        "meta_tipo": meta_tipo,
        "meta_prod": _safe_float(meta_prod_raw),
        "horas_plan": _safe_float(horas_plan_raw),
        "input_fields": input_fields,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  1. ESTRUCTURA DINÁMICA — selects encadenados Area → KPI → Campos
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/areas/stats")
def get_areas_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Retorna la lista de áreas con sus estadísticas de KPIs activos y totales
    en 1 sola petición, eliminando el efecto cascada (N+1) del frontend.
    """
    es_admin = current_user.kpi_rol_id == 1
    _cerrar_kpis_vencidos_interno(db, system_user_id=current_user.id)
    cache_key = f"kpis-areas-stats-{current_user.kpi_area_id}-{current_user.kpi_rol_id}"
    cached_data = get_cache(cache_key)
    if cached_data:
        return cached_data
    now = datetime.now()

    # 1. Traer áreas permitidas
    query_areas = db.query(Area).filter(Area.activo == True)
    if not es_admin and current_user.kpi_area_id:
        query_areas = query_areas.filter(Area.id == current_user.kpi_area_id)
    areas = query_areas.order_by(Area.nombre).all()

    # 2. Traer KPIs y programaciones activas (1 solo query cada uno)
    kpis = db.query(Kpi).filter(Kpi.activo == True).all()
    programaciones = db.query(KpiProgramado).join(Kpi).filter(
        KpiProgramado.fecha_inicio <= now,
        KpiProgramado.fecha_fin >= now
    ).all()

    # 3. Contar en memoria (Super rápido, sin saturar la DB)
    prog_area_count = {}
    for p in programaciones:
        prog_area_count[p.kpi.area_id] = prog_area_count.get(p.kpi.area_id, 0) + 1
        
    kpi_area_count = {}
    for k in kpis:
        kpi_area_count[k.area_id] = kpi_area_count.get(k.area_id, 0) + 1

    # 4. Formatear salida
    res = []
    for a in areas:
        res.append({
            "id": a.id,
            "nombre": a.nombre,
            "total": kpi_area_count.get(a.id, 0),
            "activos": prog_area_count.get(a.id, 0),
            "max": 3
        })
    set_cache(cache_key, res)
    return res


@router.get("/area/{area_id:int}")
def get_kpis_por_area(
    area_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    es_admin = current_user.kpi_rol_id == 1
    if not es_admin and current_user.kpi_area_id != area_id:
        raise HTTPException(status_code=403, detail="No tienes acceso a esta área.")

    kpis = (
        db.query(Kpi)
        .filter(Kpi.area_id == area_id, Kpi.activo == True)
        .order_by(Kpi.nombre)
        .all()
    )
    return [{"id": k.id, "nombre": k.nombre, "formula_texto": k.formula_texto} for k in kpis]


@router.get("/campos/{kpi_id:int}")
def get_campos_kpi(kpi_id: int, db: Session = Depends(get_db)):
    campos = (
        db.query(KpiCampo)
        .filter(KpiCampo.kpi_id == kpi_id)
        .order_by(KpiCampo.orden)
        .all()
    )
    kpi = db.query(Kpi).filter(Kpi.id == kpi_id).first()

    kpi_meta = None
    if kpi:
        kpi_meta = {
            "meta_valor": kpi.meta_valor,
            "meta_produccion": getattr(kpi, "meta_produccion", None),
            "horas_planificadas": getattr(kpi, "horas_planificadas", None),
        }

    campos_serializados = [
        {
            "id": c.id,
            "kpi_id": c.kpi_id,
            "campo_key": c.campo_key,
            "campo_label": c.campo_label,
            "tipo": c.tipo,
            "origen": c.origen,
            "formula_personalizada": c.formula_personalizada or "",
            "es_requerido": c.es_requerido,
            "orden": c.orden,
        }
        for c in campos
    ]

    return {"campos": campos_serializados, "kpi_meta": kpi_meta}


# ══════════════════════════════════════════════════════════════════════════════
#  2. CONFIGURACIÓN DEL MINI EXCEL
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/configuracion/{kpi_id:int}")
def get_configuracion(
    kpi_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.kpi_rol_id != 1:
        raise HTTPException(status_code=403, detail="Solo el administrador puede ver la configuración de KPIs.")

    campos = (
        db.query(KpiCampo)
        .filter(KpiCampo.kpi_id == kpi_id)
        .order_by(KpiCampo.orden)
        .all()
    )
    kpi = db.query(Kpi).filter(Kpi.id == kpi_id).first()
    if not kpi:
        raise HTTPException(status_code=404, detail="KPI no encontrado.")

    return {
        "formula_original": kpi.formula_texto or "",
        "campos": [
            {
                "id": c.id,
                "campo_key": c.campo_key,
                "campo_label": c.campo_label,
                "tipo": c.tipo,
                "origen": c.origen,
                "formula_personalizada": c.formula_personalizada or "",
            }
            for c in campos
        ],
    }


@router.post("/configuracion/{kpi_id:int}")
def save_configuracion(
    kpi_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.kpi_rol_id != 1:
        raise HTTPException(status_code=403, detail="Solo el administrador puede configurar KPIs.")

    campos_payload = payload.get("campos", [])
    for c in campos_payload:
        campo = (
            db.query(KpiCampo)
            .filter(KpiCampo.id == c["id"], KpiCampo.kpi_id == kpi_id)
            .first()
        )
        if campo:
            campo.origen = c.get("origen", campo.origen)
            campo.formula_personalizada = c.get("formula_personalizada") or None

    try:
        db.commit()
        invalidate_cache_prefix("kpis-") # <-- AGREGAR
        return {"success": True, "message": "Configuración guardada exitosamente"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ══════════════════════════════════════════════════════════════════════════════
#  3. DASHBOARD
#  FIX: Optimización N+1 resuelta en memoria mediante diccionarios para no 
#  depender de 'Area.kpis' en models.py. (1 query para áreas, 1 para KPIs).
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/dashboard_data")
def get_dashboard_data(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    es_admin = current_user.kpi_rol_id == 1

    # 1. Obtener áreas permitidas
    query_areas = db.query(Area).filter(Area.activo == True)
    if not es_admin:
        if not current_user.kpi_area_id:
            return []
        query_areas = query_areas.filter(Area.id == current_user.kpi_area_id)
        
    areas = query_areas.order_by(Area.nombre).all()
    area_ids = [a.id for a in areas]

    # 2. Obtener TODOS los KPIs de esas áreas en 1 sola consulta
    if not area_ids:
        return []
        
    todos_kpis = db.query(Kpi).filter(Kpi.area_id.in_(area_ids)).all()
    
    # 3. Agrupar en memoria (Súper rápido y elimina el N+1)
    kpis_por_area = {a_id: [] for a_id in area_ids}
    for k in todos_kpis:
        kpis_por_area[k.area_id].append(k)

    # 4. Formatear la respuesta
    result = []
    for area in areas:
        kpis_ordenados = sorted(kpis_por_area[area.id], key=lambda k: k.id)
        result.append(
            {
                "id": area.id,
                "nombre": area.nombre,
                "kpis": [
                    {
                        "id": k.id,
                        "nombre": k.nombre,
                        "formula_texto": k.formula_texto or "N/A",
                        "tipo_kpi": getattr(k, "tipo_kpi", "Positivo") or "Positivo",
                        "activo_semanal": k.activo_semanal,
                    }
                    for k in kpis_ordenados
                ],
            }
        )
    return result

# ══════════════════════════════════════════════════════════════════════════════
#  3b. KPIs semanales por área (panel de gestión semanal)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/semanales/{area_id:int}")
def get_kpis_semanales_por_area(
    area_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # ✅ FIX: El Admin (1) y el Jefe (2) pueden LEER esto. 
    # (El Jefe lo necesita en la pantalla "Mi Equipo" para poder repartirlos).
    if current_user.kpi_rol_id not in [1, 2]:
        raise HTTPException(status_code=403, detail="Sin permisos para gestionar KPIs semanales.")

    if current_user.kpi_rol_id == 2 and current_user.kpi_area_id != area_id:
        raise HTTPException(status_code=403, detail="Solo puedes gestionar tu propia área.")

    area = db.query(Area).filter(Area.id == area_id).first()
    if not area:
        raise HTTPException(status_code=404, detail="Área no encontrada.")

    kpis = db.query(Kpi).filter(Kpi.area_id == area_id, Kpi.activo == True).order_by(Kpi.nombre).all()

    _cerrar_kpis_vencidos_interno(db, system_user_id=current_user.id)

    now = datetime.now()

    programaciones = db.query(KpiProgramado).join(Kpi).filter(
        Kpi.area_id == area_id,
        KpiProgramado.fecha_inicio <= now,
        KpiProgramado.fecha_fin >= now,
    ).all()

    prog_dict = {p.kpi_id: p for p in programaciones}

    return {
        "area_id": area.id,
        "area_nombre": area.nombre,
        "activos_count": len(programaciones),
        "max_activos": 3,
        "kpis": [
            {
                "id": k.id,
                "nombre": k.nombre,
                "formula_texto": k.formula_texto or "",
                "is_programado": k.id in prog_dict,
                "fecha_fin": prog_dict[k.id].fecha_fin if k.id in prog_dict else None,
                "responsable_id": k.responsable_id,
                "completado": prog_dict[k.id].completado if k.id in prog_dict else False,
            }
            for k in kpis
        ],
    }
# ══════════════════════════════════════════════════════════════════════════════
#  4. OPERACIÓN DIARIA
#  FIX (🔴 Alto): Se elimina N+1 en serialización agregando joinedload anidado.
#  Antes: 1 query para programados + 1 query por p.kpi + 1 query por p.kpi.responsable
#         = hasta 2N queries adicionales (con N=50 → ~100 queries extra).
#  Ahora: 1 solo query con joinedload(kpi) + joinedload(kpi.responsable).
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/diario", response_model=List[KpiResponse])
def obtener_kpis_diarios(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    es_admin = current_user.kpi_rol_id == 1
    _cerrar_kpis_vencidos_interno(db, system_user_id=current_user.id)
    now = datetime.now()

    query = (
        db.query(KpiProgramado)
        .options(
            joinedload(KpiProgramado.kpi).joinedload(Kpi.responsable)
        )
        .join(Kpi)
        .filter(
            KpiProgramado.fecha_inicio <= now,
            KpiProgramado.fecha_fin >= now,
        )
    )

    if not es_admin:
        if not current_user.kpi_area_id:
            return []
        query = query.filter(Kpi.area_id == current_user.kpi_area_id)

    programados = query.all()

    # p.kpi y p.kpi.responsable ya están en memoria — sin queries adicionales
    resultado = [
        KpiResponse(
            id=p.kpi.id,
            nombre=p.kpi.nombre,
            area_id=p.kpi.area_id,
            responsable_id=p.kpi.responsable_id,
            responsable_nombre=p.kpi.responsable.name if p.kpi.responsable else None,
            meta_valor=p.kpi.meta_valor or 0.0,
            activo_semanal=True,
            es_mi_kpi=(p.kpi.responsable_id == current_user.id),
            fecha_fin=p.fecha_fin,
            completado=p.completado
        )
        for p in programados
    ]
    return sorted(resultado, key=lambda k: k.es_mi_kpi, reverse=True)

@router.post("/{kpi_id}/programar")
def programar_kpi_semanal(
    kpi_id: int,
    payload: KpiProgramar,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Programa un KPI con fecha y hora de inicio y fin en la tabla de auditoría."""
    if current_user.kpi_rol_id != 1:
        raise HTTPException(status_code=403, detail="Solo el administrador puede programar KPIs.")

    kpi = db.query(Kpi).filter(Kpi.id == kpi_id).first()
    if not kpi:
        raise HTTPException(status_code=404, detail="KPI no encontrado.")

    # 🔴 NUEVO: Buscamos el nombre exacto del área en la BD para que PHP lo entienda
    area = db.query(Area).filter(Area.id == kpi.area_id).first()
    nombre_area_texto = area.nombre if area else ""

    activos_count = db.query(KpiProgramado).join(Kpi).filter(
        Kpi.area_id == kpi.area_id,
        KpiProgramado.fecha_inicio <= payload.fecha_fin,
        KpiProgramado.fecha_fin >= payload.fecha_inicio
    ).count()

    if activos_count >= 3:
        raise HTTPException(status_code=400, detail="El área ya tiene el máximo de 3 KPIs programados en este rango de fechas.")

    nuevo_programado = KpiProgramado(
        kpi_id=kpi_id,
        fecha_inicio=payload.fecha_inicio,
        fecha_fin=payload.fecha_fin,
        asignado_por=current_user.id
    )
    db.add(nuevo_programado)
    db.commit()
    invalidate_cache_prefix("kpis-") # <-- AGREGAR


    try:
        crear_notificacion(
            db=db,
            title="🎯 Nueva Tarea de Área",
            body=f"Se ha habilitado el KPI '{kpi.nombre}' para tu área. Límite para registro: {payload.fecha_fin.strftime('%d/%m %I:%M %p')}.",
            audience="area",
            audience_value=nombre_area_texto, # 🔴 FIX: Enviamos "DESARROLLO Y PROGRAMACIÓN WEB" en vez de "17"
            created_by=current_user.id      
        )
    except Exception as e:
        print(f"Error al enviar notificación: {e}")
        
    return {"message": "KPI programado exitosamente."}


# ══════════════════════════════════════════════════════════════════════════════
#  5. REGISTRO DE VALORES
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/registrar")
def registrar_llenado(
    registro: RegistroCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    campos = (
        db.query(KpiCampo).filter(KpiCampo.kpi_id == registro.kpi_id).all()
    )

    valor_semanal = cumplimiento = productividad = eficiencia = None
    eficacia = efectividad = rendimiento = None
    alerta = "gris"
    observaciones = ""
    acciones_correctivas = ""

    for c in campos:
        key = c.campo_key
        label = c.campo_label.lower()
        if key not in registro.valores:
            continue
        val = registro.valores[key]
        if val is None or str(val).strip() == "":
            continue

        if "valor semanal" in label:
            valor_semanal = _safe_float(val)
        elif "cumplimiento" in label:
            cumplimiento = _safe_float(val)
        elif "productividad" in label:
            productividad = _safe_float(val)
        elif "eficiencia" in label:
            eficiencia = _safe_float(val)
        elif "eficacia" in label:
            eficacia = _safe_float(val)
        elif "efectividad" in label:
            efectividad = _safe_float(val)
        elif "rendimiento" in label:
            rendimiento = _safe_float(val)
        elif "alerta" in label or "semáforo" in label:
            val_str = str(val).lower()
            if "verde" in val_str:
                alerta = "verde"
            elif "amarillo" in val_str:
                alerta = "amarillo"
            elif "rojo" in val_str:
                alerta = "rojo"
        elif "observaciones" in label:
            observaciones = str(val)
        elif "acciones correctivas" in label:
            acciones_correctivas = str(val)

    nuevo_registro = RegistroKpi(
        usuario_id=current_user.id,
        kpi_id=registro.kpi_id,
        periodo_inicio=registro.periodo_inicio or datetime.now().strftime("%Y-%m-%d"),
        periodo_fin=registro.periodo_fin or datetime.now().strftime("%Y-%m-%d"),
        semana=registro.semana or datetime.now().isocalendar()[1],
        estado="enviado",
        alerta=alerta,
        valor_semanal=valor_semanal,
        cumplimiento=cumplimiento,
        productividad=productividad,
        eficiencia=eficiencia,
        eficacia=eficacia,
        efectividad=efectividad,
        rendimiento=rendimiento,
        observaciones=observaciones,
        acciones_correctivas=acciones_correctivas,
    )
    db.add(nuevo_registro)
    db.flush()
    db.refresh(nuevo_registro)

    for c in campos:
        key = c.campo_key
        if key in registro.valores:
            v_float = _safe_float(registro.valores[key])
            if v_float is not None:
                rv = RegistroValores(
                    registro_id=nuevo_registro.id, campo_id=c.id, valor=v_float
                )
                db.add(rv)

    now = datetime.now()
    programado = db.query(KpiProgramado).filter(
        KpiProgramado.kpi_id == registro.kpi_id,
        KpiProgramado.fecha_inicio <= now,
        KpiProgramado.fecha_fin >= now,
        KpiProgramado.completado == False
    ).first()

    if programado:
        programado.completado = True
        programado.registro_kpi_id = nuevo_registro.id

    db.commit()
    invalidate_cache_prefix("kpis-") # <-- AGREGAR
    return {
        "message": "Valor registrado con éxito",
        "registro_id": nuevo_registro.id,
        "alerta": alerta,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  6. GESTIÓN DE ÁREAS Y KPIs (CRUD admin)
#  FIX (🟡 Medio): Se eliminan las materializaciones innecesarias de ORM.
#  Antes: .all() cargaba objetos completos a RAM solo para extraer su .id
#  Ahora: se extrae directamente la columna ID con db.query(Model.id).filter(...)
#         y se usa .scalar_subquery() para el DELETE en cascada, sin tocar RAM.
# ══════════════════════════════════════════════════════════════════════════════

@router.delete("/areas/{area_id:int}")
def delete_area(
    area_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.kpi_rol_id != 1:
        raise HTTPException(status_code=403, detail="Solo el administrador puede eliminar áreas.")

    area = db.query(Area).filter(Area.id == area_id).first()
    if not area:
        raise HTTPException(status_code=404, detail="Área no encontrada.")

    try:
        # Subquery: IDs de KPIs del área (no materializa objetos ORM)
        kpis_subq = (
            db.query(Kpi.id)
            .filter(Kpi.area_id == area_id)
            .scalar_subquery()
        )

        # Subquery: IDs de registros asociados a esos KPIs
        registros_subq = (
            db.query(RegistroKpi.id)
            .filter(RegistroKpi.kpi_id.in_(kpis_subq))
            .scalar_subquery()
        )

        # Borrado en cascada usando subqueries — solo opera en BD, sin cargar RAM
        db.query(RegistroValores).filter(
            RegistroValores.registro_id.in_(registros_subq)
        ).delete(synchronize_session=False)

        db.query(RegistroKpi).filter(
            RegistroKpi.kpi_id.in_(kpis_subq)
        ).delete(synchronize_session=False)

        db.query(KpiCampo).filter(
            KpiCampo.kpi_id.in_(kpis_subq)
        ).delete(synchronize_session=False)

        db.query(Kpi).filter(
            Kpi.area_id == area_id
        ).delete(synchronize_session=False)

        db.delete(area)
        db.commit()
        invalidate_cache_prefix("kpis-") # <-- AGREGAR
        return {"success": True, "message": "Área y todos sus KPIs eliminados correctamente."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/kpi/{kpi_id:int}")
def delete_kpi(
    kpi_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.kpi_rol_id != 1:
        raise HTTPException(status_code=403, detail="Solo el administrador puede eliminar KPIs.")

    kpi = db.query(Kpi).filter(Kpi.id == kpi_id).first()
    if not kpi:
        raise HTTPException(status_code=404, detail="KPI no encontrado.")

    try:
        # Subquery: IDs de registros del KPI (no materializa objetos ORM)
        registros_subq = (
            db.query(RegistroKpi.id)
            .filter(RegistroKpi.kpi_id == kpi_id)
            .scalar_subquery()
        )

        # Borrado en cascada usando subqueries — solo opera en BD, sin cargar RAM
        db.query(RegistroValores).filter(
            RegistroValores.registro_id.in_(registros_subq)
        ).delete(synchronize_session=False)

        db.query(RegistroKpi).filter(
            RegistroKpi.kpi_id == kpi_id
        ).delete(synchronize_session=False)

        db.query(KpiCampo).filter(
            KpiCampo.kpi_id == kpi_id
        ).delete(synchronize_session=False)

        db.delete(kpi)
        db.commit()
        return {"success": True, "message": "KPI eliminado correctamente."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ══════════════════════════════════════════════════════════════════════════════
#  7. IMPORTACIÓN DE EXCEL
# ══════════════════════════════════════════════════════════════════════════════

@router.patch("/{kpi_id}/responsable")
def asignar_responsable_kpi(
    kpi_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.kpi_rol_id not in [1, 2]:
        raise HTTPException(status_code=403, detail="Sin permisos para asignar responsables.")

    kpi = db.query(Kpi).filter(Kpi.id == kpi_id).first()
    if not kpi:
        raise HTTPException(status_code=404, detail="KPI no encontrado.")

    if current_user.kpi_rol_id == 2 and kpi.area_id != current_user.kpi_area_id:
        raise HTTPException(status_code=403, detail="Solo puedes asignar responsables en tu área.")

    responsable_id = payload.get("responsable_id")

    if responsable_id is None:
        kpi.responsable_id = None
        db.commit()
        return {"success": True, "message": f"KPI '{kpi.nombre}' desasignado."}

    trabajador = db.query(User).filter(User.id == responsable_id).first()
    if not trabajador:
        raise HTTPException(status_code=404, detail="Trabajador no encontrado.")

    if trabajador.kpi_rol_id != 1 and trabajador.kpi_area_id != kpi.area_id:
        raise HTTPException(status_code=400, detail="El trabajador no pertenece al área de este KPI.")

    kpi.responsable_id = responsable_id
    db.commit()
    invalidate_cache_prefix("kpis-") # <-- AGREGAR
    return {"success": True, "message": f"Responsable asignado correctamente al KPI '{kpi.nombre}'."}

@router.post("/upload")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.kpi_rol_id != 1:
        raise HTTPException(status_code=403, detail="Solo el administrador puede importar Excels.")

    import tempfile
    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        result = _parse_excel_and_save(tmp_path, db)
        return {"success": True, "result": result}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        os.unlink(tmp_path)


@router.post("/upload_smart")
async def upload_smart_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.kpi_rol_id != 1:
        raise HTTPException(status_code=403, detail="Solo el administrador puede importar Excels.")

    import tempfile
    suffix = os.path.splitext(file.filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        kpis_actualizados = _process_smart_excel(tmp_path, db)
        return {
            "success": True,
            "message": f"¡Diccionario SMART procesado con éxito! Se auto-configuraron {kpis_actualizados} KPIs.",
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        os.unlink(tmp_path)


# ──────────────────────────────────────────────────────────────────────────────
#  Helpers internos de importación
# ──────────────────────────────────────────────────────────────────────────────

def _parse_excel_and_save(filepath: str, db: Session) -> dict:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    ws_inicio = wb["Inicio"]
    area_name = None
    for row in ws_inicio.iter_rows(min_row=2, max_row=5, values_only=True):
        if row[1] is not None and str(row[1]).strip():
            area_name = str(row[1]).strip()
            break
    area_name = area_name or "Sin nombre"

    area = db.query(Area).filter(Area.nombre == area_name).first()
    if not area:
        area = Area(nombre=area_name, activo=True)
        db.add(area)
        db.flush()

    SKIP_SHEETS = {
        "Inicio", "Índice de formulas", "Índice de fórmulas",
        "Leyenda de Evaluación de Indica", "Datos de la hoja", "datos de la hoja",
    }

    results = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        info = _extract_kpi_fields_from_sheet(ws)
        if not info:
            continue

        kpi_name = info["kpi_name"] or sheet_name

        kpi = Kpi(
            nombre=kpi_name,
            formula_texto=info["formula"],
            area_id=area.id,
            meta_valor=info["meta_valor"],
            meta_produccion=info["meta_prod"],
            horas_planificadas=info["horas_plan"],
            activo=True,
            activo_semanal=False,
        )
        db.add(kpi)
        db.flush()

        formula_valor_semanal = build_js_formula(info["formula"], info["input_fields"])
        campos_guardados = []

        for orden, col_label in enumerate(info["input_fields"]):
            campo_key = _campo_key(col_label)
            lower_label = col_label.lower()

            tipo = "texto" if any(
                w in lower_label for w in ["observaciones", "acciones", "fecha", "alerta"]
            ) else "numero"

            origen = "usuario"
            formula_pers = None

            if "eficiencia" in lower_label:
                origen, formula_pers = "calculado", "([Horas planificadas] / [Horas reales])"
            elif "eficacia" in lower_label:
                origen, formula_pers = "calculado", "[Cumplimiento (%)] > 1 ? 1 : [Cumplimiento (%)]"
            elif "efectividad" in lower_label:
                origen, formula_pers = "calculado", "([Eficiencia (%)] * [Eficacia (%)])"
            elif "rendimiento" in lower_label:
                origen, formula_pers = "calculado", "([Productividad] / [Meta Producción ≥])"
            elif "alerta" in lower_label:
                origen = "sistema"
            elif "valor semanal" in lower_label:
                origen, formula_pers = "calculado", formula_valor_semanal
            elif any(w in lower_label for w in ["cumplimiento", "productividad", "calculado"]):
                origen = "calculado"
            elif "meta" in lower_label:
                origen = "sistema"

            existing = (
                db.query(KpiCampo)
                .filter(KpiCampo.kpi_id == kpi.id, KpiCampo.campo_key == campo_key)
                .first()
            )
            if not existing:
                campo = KpiCampo(
                    kpi_id=kpi.id,
                    campo_key=campo_key,
                    campo_label=col_label,
                    tipo=tipo,
                    origen=origen,
                    formula_personalizada=formula_pers,
                    es_requerido=False,
                    orden=orden,
                )
                db.add(campo)
                campos_guardados.append(col_label)

        results.append({"sheet": sheet_name, "kpi": kpi_name, "campos_entrada": campos_guardados})

    db.commit()
    wb.close()
    return {"area": area_name, "kpis": results}


def _process_smart_excel(filepath: str, db: Session) -> int:
    df = pd.read_excel(filepath, sheet_name="2_Metas")
    df = df.fillna("")
    kpis_actualizados = 0

    for _, row in df.iterrows():
        nombre_kpi = str(row.get("Nombre KPI", "")).strip()
        tipo_kpi_crudo = str(row.get("Tipo_KPI", "")).strip().lower()
        if not nombre_kpi:
            continue

        kpi = db.query(Kpi).filter(Kpi.nombre == nombre_kpi).first()
        if not kpi:
            continue

        tipo_enum = "Negativo" if tipo_kpi_crudo == "negativo" else "Positivo"
        if hasattr(kpi, "tipo_kpi"):
            kpi.tipo_kpi = tipo_enum

        campos = db.query(KpiCampo).filter(KpiCampo.kpi_id == kpi.id).all()

        label_valor = "[Valor semanal]"
        label_meta = "[Meta KPI]"
        label_meta_prod = "[Meta Producción]"

        for c in campos:
            lbl = c.campo_label.lower()
            if "valor semanal" in lbl:
                label_valor = f"[{c.campo_label}]"
            elif "meta kpi" in lbl:
                label_meta = f"[{c.campo_label}]"
            elif "meta producción" in lbl or "meta produccion" in lbl:
                label_meta_prod = f"[{c.campo_label}]"

        if tipo_enum == "Positivo":
            f_cump = f"({label_meta} === 0 || {label_meta} === null) ? 0 : ({label_valor} / {label_meta})"
            f_prod = f"({label_meta_prod} === 0 || {label_meta_prod} === null) ? 0 : ({label_valor} / {label_meta_prod})"
        else:
            f_cump = (
                f"({label_meta} === 0 || {label_meta} === null) ? "
                f"({label_valor} === 0 ? 1 : 0) : Math.max(0, 1 - ({label_valor} / {label_meta}))"
            )
            f_prod = (
                f"({label_meta_prod} === 0 || {label_meta_prod} === null) ? "
                f"({label_valor} === 0 ? 1 : 0) : Math.max(0, 1 - ({label_valor} / {label_meta_prod}))"
            )

        for c in campos:
            lbl = c.campo_label.lower()
            if "cumplimiento" in lbl:
                c.formula_personalizada = f_cump
                c.origen = "calculado"
            elif "productividad" in lbl:
                c.formula_personalizada = f_prod
                c.origen = "calculado"

        kpis_actualizados += 1

    db.commit()
    return kpis_actualizados


# ══════════════════════════════════════════════════════════════════════════════
#  8. CIERRE AUTOMÁTICO DE KPIs — FUNCIÓN INTERNA + ENDPOINT MANUAL
#  FIX (🔴 Alto): Se elimina el N+1 implícito en el loop.
#  Antes: p.kpi.responsable_id dentro del loop disparaba 1 query extra por
#         cada KPI vencido para cargar la relación lazy `kpi` (y potencialmente
#         otra para `kpi.responsable`). Con 50 vencidos = ~100 queries extra.
#  Ahora: joinedload(KpiProgramado.kpi) trae toda la relación en 1 solo query.
# ══════════════════════════════════════════════════════════════════════════════

_last_cerrar_kpis_ts = {"value": 0}

def _cerrar_kpis_vencidos_interno(db: Session, system_user_id: str | None = None) -> int:
    now_ts = _time.time()
    if now_ts - _last_cerrar_kpis_ts["value"] < 300:
        return 0
    _last_cerrar_kpis_ts["value"] = now_ts

    now = datetime.now()

    # joinedload precarga kpi (y su responsable_id) en el mismo query inicial
    vencidos = (
        db.query(KpiProgramado)
        .options(joinedload(KpiProgramado.kpi))
        .filter(
            KpiProgramado.fecha_fin < now,
            KpiProgramado.completado == False,
        )
        .all()
    )

    if not vencidos:
        return 0

    cerrados_count = 0
    for p in vencidos:
        # p.kpi ya está en memoria — sin query adicional
        responsable_omision = p.kpi.responsable_id or p.asignado_por or system_user_id

        if not responsable_omision:
            continue

        registro_omision = RegistroKpi(
            usuario_id=responsable_omision,
            kpi_id=p.kpi_id,
            periodo_inicio=p.fecha_inicio.strftime("%Y-%m-%d"),
            periodo_fin=p.fecha_fin.strftime("%Y-%m-%d"),
            semana=p.fecha_inicio.isocalendar()[1],
            estado="no_reportado",
            alerta="rojo",
            valor_semanal=0.0,
            observaciones="Cierre automático por sistema (fecha vencida sin llenado)",
        )
        db.add(registro_omision)
        db.flush()

        p.completado = True
        p.registro_kpi_id = registro_omision.id
        cerrados_count += 1

    if cerrados_count:
        db.commit()
        invalidate_cache_prefix("kpis-") # <-- AGREGAR (Para que limpie el caché si el sistema cerró algo por su cuenta)

    return cerrados_count


@router.post("/cerrar-vencidos")
def cerrar_kpis_vencidos(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Endpoint manual para forzar el cierre. En operación normal ocurre automáticamente."""
    if current_user.kpi_rol_id != 1:
        raise HTTPException(status_code=403, detail="Solo el administrador puede ejecutar el cierre de KPIs.")

    cerrados = _cerrar_kpis_vencidos_interno(db, system_user_id=current_user.id)
    if cerrados == 0:
        return {"success": True, "message": "No hay KPIs vencidos pendientes por cerrar.", "cerrados": 0}

    return {
        "success": True,
        "message": f"Se cerraron {cerrados} KPIs vencidos con reporte de omisión (Rojo).",
        "cerrados": cerrados,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  9. PANEL DE ALERTAS Y REPORTES (MÓDULO 3)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/alertas")
def obtener_alertas_dashboard(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.kpi_rol_id not in [1, 2]:
        raise HTTPException(status_code=403, detail="Sin permisos para ver el panel de alertas.")

    _cerrar_kpis_vencidos_interno(db, system_user_id=current_user.id)

    es_admin = current_user.kpi_rol_id == 1
    filtro_area = current_user.kpi_area_id if not es_admin else None

    cache_key = f"kpis-alertas-{filtro_area}"
    cached_data = get_cache(cache_key)
    if cached_data:
        return cached_data

    now = datetime.now()
    es_admin = current_user.kpi_rol_id == 1
    filtro_area = current_user.kpi_area_id if not es_admin else None

    # -------------------------------------------------------------------------
    # BLOQUE 1: PENDIENTES DE LLENADO
    # -------------------------------------------------------------------------
    query_pendientes = db.query(KpiProgramado, Kpi, User, Area).join(Kpi, KpiProgramado.kpi_id == Kpi.id)\
        .outerjoin(User, Kpi.responsable_id == User.id)\
        .join(Area, Kpi.area_id == Area.id)\
        .filter(
            KpiProgramado.fecha_inicio <= now,
            KpiProgramado.fecha_fin >= now,
            KpiProgramado.completado == False
        )

    if filtro_area:
        query_pendientes = query_pendientes.filter(Kpi.area_id == filtro_area)

    pendientes_raw = query_pendientes.all()
    pendientes = []
    for p, k, u, a in pendientes_raw:
        pendientes.append({
            "kpi_nombre": k.nombre,
            "area_nombre": a.nombre,
            "responsable": u.name if u else "Sin asignar",
            "fecha_fin": p.fecha_fin,
            "dias_restantes": (p.fecha_fin - now).days
        })

    # -------------------------------------------------------------------------
    # BLOQUE 2: REGISTROS EN RIESGO
    # -------------------------------------------------------------------------
    query_riesgo = db.query(RegistroKpi, Kpi, User, Area).join(Kpi, RegistroKpi.kpi_id == Kpi.id)\
        .join(User, RegistroKpi.usuario_id == User.id)\
        .join(Area, Kpi.area_id == Area.id)\
        .filter(RegistroKpi.alerta.in_(["rojo", "amarillo"]))\
        .order_by(RegistroKpi.id.desc())

    if filtro_area:
        query_riesgo = query_riesgo.filter(Kpi.area_id == filtro_area)

    riesgo_raw = query_riesgo.limit(10).all()
    en_riesgo = []
    for r, k, u, a in riesgo_raw:
        en_riesgo.append({
            "id_registro": r.id,
            "kpi_nombre": k.nombre,
            "area_nombre": a.nombre,
            "responsable": u.name,
            "alerta": r.alerta,
            "valor_registrado": r.valor_semanal,
            "estado": r.estado,
            "fecha": r.periodo_fin
        })

    # -------------------------------------------------------------------------
    # BLOQUE 3: PARTICIPACIÓN
    # -------------------------------------------------------------------------
    query_participacion = db.query(KpiProgramado, Kpi, Area).join(Kpi, KpiProgramado.kpi_id == Kpi.id)\
        .join(Area, Kpi.area_id == Area.id)\
        .filter(
            KpiProgramado.fecha_inicio <= now,
            KpiProgramado.fecha_fin >= now
        )

    if filtro_area:
        query_participacion = query_participacion.filter(Kpi.area_id == filtro_area)

    part_raw = query_participacion.all()

    stats_por_area = {}
    for p, k, a in part_raw:
        if a.nombre not in stats_por_area:
            stats_por_area[a.nombre] = {"total": 0, "completados": 0}
        stats_por_area[a.nombre]["total"] += 1
        if p.completado:
            stats_por_area[a.nombre]["completados"] += 1

    participacion = []
    for area_nom, stats in stats_por_area.items():
        porcentaje = (stats["completados"] / stats["total"] * 100) if stats["total"] > 0 else 0
        participacion.append({
            "area": area_nom,
            "total_programados": stats["total"],
            "completados": stats["completados"],
            "porcentaje": round(porcentaje, 1)
        })

    resultado = {
        "pendientes": pendientes,
        "en_riesgo": en_riesgo,
        "participacion": participacion
    }
    set_cache(cache_key, resultado)
    return resultado


# ══════════════════════════════════════════════════════════════════════════════
#  10. MIS REPORTES — Historial personal (todos los roles)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/mis-reportes")
def obtener_mis_reportes(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cache_key = f"kpis-mis-reportes-{current_user.id}"
    cached_data = get_cache(cache_key)
    if cached_data:
        return cached_data

    registros = (
        db.query(RegistroKpi, Kpi)
        .join(Kpi, RegistroKpi.kpi_id == Kpi.id)
        .filter(RegistroKpi.usuario_id == current_user.id)
        .order_by(RegistroKpi.id.desc())
        .all()
    )

    resultado = [
        {
            "id": r.id,
            "kpi_nombre": k.nombre,
            "periodo_inicio": r.periodo_inicio,
            "periodo_fin": r.periodo_fin,
            "estado": r.estado,
            "valor_semanal": r.valor_semanal,
            "cumplimiento": r.cumplimiento,
            "alerta": r.alerta,
            "observaciones": r.observaciones,
            "enviado_en": r.enviado_en,
        }
        for r, k in registros
    ]
    set_cache(cache_key, resultado)
    return resultado


# ══════════════════════════════════════════════════════════════════════════════
#  11. HISTORIAL GENERAL — Solo Admin (Rol 1)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/historial")
def obtener_historial_general(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.kpi_rol_id != 1:
        raise HTTPException(status_code=403, detail="Solo el administrador puede ver el historial general.")

    cache_key = "kpis-historial"
    cached_data = get_cache(cache_key)
    if cached_data:
        return cached_data

    registros = (
        db.query(RegistroKpi, Kpi, User, Area)
        .join(Kpi, RegistroKpi.kpi_id == Kpi.id)
        .join(User, RegistroKpi.usuario_id == User.id)
        .join(Area, Kpi.area_id == Area.id)
        .order_by(RegistroKpi.id.desc())
        .all()
    )

    resultado = [
        {
            "id": r.id,
            "periodo_inicio": r.periodo_inicio,
            "periodo_fin": r.periodo_fin,
            "area_nombre": a.nombre,
            "kpi_nombre": k.nombre,
            "responsable": u.name,
            "valor_semanal": r.valor_semanal,
            "cumplimiento": r.cumplimiento,
            "productividad": r.productividad,
            "eficiencia": r.eficiencia,
            "eficacia": r.eficacia,
            "efectividad": r.efectividad,
            "rendimiento": r.rendimiento,
            "estado": r.estado,
            "alerta": r.alerta,
            "observaciones": r.observaciones,
            "acciones_correctivas": r.acciones_correctivas,
            "enviado_en": r.enviado_en,
        }
        for r, k, u, a in registros
    ]
    set_cache(cache_key, resultado)
    return resultado